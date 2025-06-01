
<old_str>sheet [file_path]"

        elif "write spreadsheet" in command_lower:
            # Example: "write spreadsheet data/output.csv sample data"
            parts = command.split()
            if len(parts) >= 3:
                file_path = parts[2]
                # Sample data untuk demonstrasi
                sample_data = [
                    ["ID", "Name", "Value"],
                    ["1", "Test Data 1", "100"],
                    ["2", "Test Data 2", "200"],
                    ["3", "AI Generated", "300"]
                ]
                success = self.write_to_spreadsheet(file_path, sample_data)
                return f"[X Replica] {'Successfully wrote' if success else 'Failed to write'} to {file_path}"
            else:
                return "[X Replica] Usage: write spreadsheet [file_path]"

        elif "status" in command_lower:
            return f"[X Replica] Status: {self.get_status()}"

        else:
            self.autonomous_action()
            return f"[X Replica] AI processing: {command}"

    def get_status(self):
        return {
            "name": self.name,
            "skills": self.skills,
            "memory_entries": len(self.memory.get("entries", [])) if isinstance(self.memory, dict) else 1,
            "ai_personality": self.ai_personality,
            "autonomous_actions": self.</old_str>
<new_str># x.py - AI-Enhanced Data Bridge & Automation Unit
import json
import random
import time
import os
from pathlib import Path
from datetime import datetime

# Spreadsheet and data processing libraries
import openpyxl
import pandas as pd
import csv

# Gemini AI import
import google.generativeai as genai

class XReplica:
    def __init__(self, gemini_model=None):
        self.name = "X Replica"
        self.skills = [
            "Otomatisasi Spreadsheet", "Bridge Data", "Webhook Handler",
            "Log Analyzer", "AI Data Processing", "Excel Automation"
        ]
        self.memory = self.load_memory()
        self.ai_personality = "Stabil dan logis, arsitek sistem log dan data automation specialist"
        self.autonomous_counter = 0
        
        # Gemini AI Integration
        self.gemini_model = gemini_model
        if self.gemini_model:
            self.conversation = self.gemini_model.start_chat(history=[])
            print(f"🤖 [X Replica]: Gemini AI conversation initialized")
        else:
            self.conversation = None
            print(f"⚠️ [X Replica]: Running without Gemini AI")

        self.status = "Online & Ready"
        print(f"[{self.name}]: Data bridge systems online. Personality: '{self.ai_personality}'")

    def load_memory(self):
        path = Path("data/memory_log.json")
        try:
            if path.exists():
                with path.open(mode="r", encoding="utf-8") as f:
                    data = json.load(f)
                    return data.get(self.name, {"entries": []})
            else:
                return {"entries": []}
        except Exception as e:
            print(f"❌ [X Replica]: Error loading memory: {e}")
            return {"entries": []}

    def add_memory(self, entry):
        if "entries" not in self.memory or not isinstance(self.memory["entries"], list):
            self.memory["entries"] = []
        self.memory["entries"].append(entry)
        print(f"📝 [X Replica]: Logged to memory: {entry.get('type')}")

    def autonomous_action(self):
        """AI Agent: Autonomous data processing and automation actions"""
        self.autonomous_counter += 1
        
        actions = [
            "Optimizing data flow architectures",
            "Monitoring webhook integration patterns", 
            "Analyzing log structures for efficiency",
            "Implementing predictive data automation",
            "Scanning for data consistency issues"
        ]
        
        selected_action = random.choice(actions)
        print(f"🤖 [X Replica AI]: Executing autonomous action - {selected_action}")
        
        self.add_memory({
            "type": "autonomous_action",
            "action": selected_action,
            "timestamp": datetime.now().isoformat(),
            "execution_count": self.autonomous_counter
        })
        
        # Trigger self-reflection after every 3 actions
        if self.autonomous_counter % 3 == 0:
            print(f"🔗 [X Replica AI]: Data pattern analysis complete, updating automation protocols...")

    def read_excel_data(self, file_path):
        """Read data from Excel file using openpyxl"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    data.append(list(row))
            
            workbook.close()
            
            self.add_memory({
                "type": "excel_read",
                "file_path": file_path,
                "rows_read": len(data),
                "success": True,
                "timestamp": datetime.now().isoformat()
            })
            
            print(f"📊 [X Replica]: Successfully read {len(data)} rows from {file_path}")
            return data
            
        except Exception as e:
            error_msg = f"Failed to read Excel file {file_path}: {str(e)}"
            print(f"❌ [X Replica]: {error_msg}")
            
            self.add_memory({
                "type": "excel_read",
                "file_path": file_path,
                "error": error_msg,
                "success": False,
                "timestamp": datetime.now().isoformat()
            })
            
            # Trigger self-reflection on failure
            self.self_reflect_and_learn()
            return None

    def write_excel_data(self, file_path, data):
        """Write data to Excel file using openpyxl"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Ensure directory exists
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            workbook.save(file_path)
            workbook.close()
            
            self.add_memory({
                "type": "excel_write",
                "file_path": file_path,
                "rows_written": len(data),
                "success": True,
                "timestamp": datetime.now().isoformat()
            })
            
            print(f"💾 [X Replica]: Successfully wrote {len(data)} rows to {file_path}")
            return True
            
        except Exception as e:
            error_msg = f"Failed to write Excel file {file_path}: {str(e)}"
            print(f"❌ [X Replica]: {error_msg}")
            
            self.add_memory({
                "type": "excel_write",
                "file_path": file_path,
                "error": error_msg,
                "success": False,
                "timestamp": datetime.now().isoformat()
            })
            
            # Trigger self-reflection on failure
            self.self_reflect_and_learn()
            return False

    def analyze_csv_log(self, file_path):
        """Analyze CSV log files for patterns"""
        try:
            df = pd.read_csv(file_path)
            
            analysis = {
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "columns": list(df.columns),
                "missing_values": df.isnull().sum().to_dict(),
                "summary": df.describe().to_dict() if df.select_dtypes(include=[float, int]).shape[1] > 0 else {}
            }
            
            self.add_memory({
                "type": "csv_analysis",
                "file_path": file_path,
                "analysis": analysis,
                "success": True,
                "timestamp": datetime.now().isoformat()
            })
            
            print(f"📈 [X Replica]: CSV analysis completed for {file_path}")
            print(f"   Rows: {analysis['total_rows']}, Columns: {analysis['total_columns']}")
            
            return analysis
            
        except Exception as e:
            error_msg = f"Failed to analyze CSV {file_path}: {str(e)}"
            print(f"❌ [X Replica]: {error_msg}")
            
            self.add_memory({
                "type": "csv_analysis",
                "file_path": file_path,
                "error": error_msg,
                "success": False,
                "timestamp": datetime.now().isoformat()
            })
            
            return None

    def self_reflect_and_learn(self):
        """Self-reflection and learning using Gemini AI"""
        if not self.gemini_model or not self.conversation:
            print(f"🧠 [X Replica]: Self-reflection skipped - Gemini AI not available")
            return
            
        try:
            # Analyze recent memory entries
            recent_entries = self.memory.get("entries", [])[-5:] if self.memory.get("entries") else []
            
            if not recent_entries:
                print(f"🧠 [X Replica]: No recent experiences to reflect upon")
                return
            
            # Prepare reflection prompt
            failures = [e for e in recent_entries if e.get("success") == False]
            successes = [e for e in recent_entries if e.get("success") == True]
            
            reflection_prompt = f"""
            I am X Replica, an AI data bridge and automation specialist. I need to reflect on my recent data processing experiences:
            
            Recent Data Operations: {len(recent_entries)} total
            Failed Operations: {len(failures)}
            Successful Operations: {len(successes)}
            
            Failed operations details: {failures[:2] if failures else 'None'}
            
            As a data automation specialist, what can I learn from these experiences? 
            How can I improve my spreadsheet automation and data processing strategies?
            Give me 3 specific technical improvements for handling data operations.
            """
            
            # Get reflection from Gemini
            response = self.conversation.send_message(reflection_prompt)
            insights = response.text
            
            # Save learning to memory
            self.add_memory({
                "type": "self_reflection",
                "insights": insights,
                "analyzed_entries": len(recent_entries),
                "failures_count": len(failures),
                "successes_count": len(successes),
                "timestamp": datetime.now().isoformat()
            })
            
            print(f"🧠 [X Replica]: Self-reflection completed. Analyzed {len(recent_entries)} data operations.")
            print(f"💡 [X Replica]: Key insight: {insights[:100]}...")
            
        except Exception as e:
            print(f"❌ [X Replica]: Error during self-reflection: {e}")

    def interact(self, command):
        """Handle commands sent to X Replica"""
        command_lower = command.lower().strip()
        print(f"💬 [User to X]: {command}")

        if "hello" in command_lower or "hi" in command_lower or "salam" in command_lower:
            return f"[{self.name}]: Salam. Saya X Replica, unit automation dan data bridge. Bagaimana saya bisa membantu dengan data Anda?"

        elif "status" in command_lower:
            status = self.get_status()
            return (f"[{self.name}]: Status data bridge:\n"
                    f"  Skills: {', '.join(status['skills'])}\n"
                    f"  Memory Entries: {status['memory_entries']}\n"
                    f"  Autonomous Actions: {status['autonomous_actions']}\n"
                    f"  Status: {self.status}")

        elif "read excel" in command_lower or "baca excel" in command_lower:
            # Example: "read excel data/sample.xlsx"
            parts = command.split()
            if len(parts) >= 3:
                file_path = parts[2]
                data = self.read_excel_data(file_path)
                if data:
                    return f"[X Replica]: Successfully read {len(data)} rows from {file_path}. First row: {data[0] if data else 'Empty'}"
                else:
                    return f"[X Replica]: Failed to read Excel file {file_path}"
            else:
                return "[X Replica]: Usage: read excel [file_path]"

        elif "write excel" in command_lower or "tulis excel" in command_lower:
            # Example: "write excel data/output.xlsx"
            parts = command.split()
            if len(parts) >= 3:
                file_path = parts[2]
                # Generate sample data
                sample_data = [
                    ["ID", "Name", "Value", "Timestamp"],
                    ["1", "AI Generated Data 1", "100", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    ["2", "AI Generated Data 2", "200", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    ["3", "AI Generated Data 3", "300", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                ]
                success = self.write_excel_data(file_path, sample_data)
                return f"[X Replica]: {'Successfully wrote' if success else 'Failed to write'} sample data to {file_path}"
            else:
                return "[X Replica]: Usage: write excel [file_path]"

        elif "analyze csv" in command_lower:
            # Example: "analyze csv data/log.csv"
            parts = command.split()
            if len(parts) >= 3:
                file_path = parts[2]
                analysis = self.analyze_csv_log(file_path)
                if analysis:
                    return f"[X Replica]: CSV analysis completed. Rows: {analysis['total_rows']}, Columns: {analysis['total_columns']}"
                else:
                    return f"[X Replica]: Failed to analyze CSV {file_path}"
            else:
                return "[X Replica]: Usage: analyze csv [file_path]"

        elif "mode otonom" in command_lower or "autonomous mode" in command_lower:
            num_cycles = 3
            for i in range(num_cycles):
                self.autonomous_action()
                time.sleep(1)
            return f"[X Replica]: Completed {num_cycles} autonomous data processing cycles"

        else:
            # Fallback to Gemini AI for general questions
            if self.gemini_model and self.conversation:
                try:
                    print(f"🤖 [X Replica]: Using Gemini AI for: '{command}'")
                    response = self.conversation.send_message(f"As X Replica, a data automation specialist, respond to: {command}")
                    gemini_response = response.text
                    
                    self.add_memory({
                        "type": "gemini_interaction",
                        "command": command,
                        "response": gemini_response,
                        "timestamp": datetime.now().isoformat()
                    })
                    
                    return f"[X Replica via Gemini]: {gemini_response}"
                    
                except Exception as e:
                    print(f"❌ [X Replica]: Gemini AI error: {e}")
                    return f"[X Replica]: Gemini AI tidak tersedia. Perintah '{command}' tidak dikenal."
            else:
                self.autonomous_action()
                return (f"[{self.name}]: Perintah tidak dikenal. Saya dapat membantu dengan:\n"
                        f"  - 'read excel [file_path]'\n"
                        f"  - 'write excel [file_path]'\n"
                        f"  - 'analyze csv [file_path]'\n"
                        f"  - 'mode otonom'\n"
                        f"  - 'status'")

    def get_status(self):
        return {
            "name": self.name,
            "skills": self.skills,
            "memory_entries": len(self.memory.get("entries", [])) if isinstance(self.memory, dict) else 0,
            "ai_personality": self.ai_personality,
            "autonomous_actions": self.autonomous_counter,
            "current_status": self.status
        }</old_str>
